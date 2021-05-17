import numpy as np
from openpyxl import Workbook
import math
import os
import sys

x_min = 324600  # minimum x value
y_min = 4150520  # minimum y value


def angle_diff(Tx, Rx, azi, dtilt):
    ln1 = Rx[0] - Tx[0]
    ln2 = Rx[1] - Tx[1]
    ln3 = Tx[2] - Rx[2]
    dist2d = math.sqrt(math.pow(ln1, 2) + math.pow(ln2, 2))

    theta = math.atan2(ln2, ln1)
    theta = theta * 180 / math.pi

    theta2 = math.atan2(ln3, dist2d)
    theta2 = theta2 * 180 / math.pi
    case = 0
    if ln2 > 0:
        if ln1 > 0:
            theta = 90 - theta
            case = 1
        else:
            theta = 450 - theta
            case = 2
    else:
        theta = 90 + theta * (-1)
        case = 3
    azi_diff = abs(theta - azi)
    tilt_diff = abs(theta2 - dtilt)

    alpha = ln3 / math.tan(dtilt)
    gamma = math.sqrt(math.pow(alpha, 2) + math.pow(ln3, 2))
    beta = math.sqrt(
        math.pow(alpha * math.sin(azi_diff), 2)
        + math.pow(dist2d - alpha * math.cos(azi_diff), 2)
    )
    dist3d = math.sqrt(math.pow(ln1, 2) + math.pow(ln2, 2) + math.pow(ln3, 2))
    f_diff = math.acos(
        (math.pow(dist3d, 2) + math.pow(gamma, 2) - math.pow(beta, 2))
        / (2 * dist3d * gamma)
    )
    # f_diff = (math.pow(dist3d, 2) + math.pow(gamma, 2) - math.pow(beta, 2)) / (2 * dist3d * gamma)
    f_diff = f_diff / math.pi * 180
    return case, azi_diff, tilt_diff, dist3d, azi_diff * dist3d


def load_npy():
    dim_1 = np.load("project/static/maps/Building.npy")
    dim_2 = np.load("project/static/maps/Terrain.npy")

    return dim_1, dim_2


def main():
    dim_1, dim_2 = load_npy()
    cnt = 1
    write_wb = Workbook()
    write_ws = write_wb.active
    write_ws["A1"] = "idx"
    write_ws["B1"] = "RX"
    write_ws["C1"] = "RY"
    write_ws["D1"] = "RZ"
    write_ws["E1"] = "TX"
    write_ws["F1"] = "TY"
    write_ws["G1"] = "TZ"
    write_ws["H1"] = "chai1"
    write_ws["I1"] = "chai2"
    write_ws["J1"] = "dist"
    id = sys.argv[1]
    TX = sys.argv[2]
    TY = sys.argv[3]
    azimuth = sys.argv[4]
    downtilt = sys.argv[5]

    idx = 2

    for y in range(int(TY) - 200, int(TY) + 200, 5):
        for x in range(int(TX) - 200, int(TX) + 200, 5):
            if dim_1[y - y_min][x - x_min] == 0:
                write_ws.cell(idx, 1, idx - 1)
                write_ws.cell(idx, 2, x)
                write_ws.cell(idx, 3, y)
                write_ws.cell(idx, 4, dim_1[y - y_min][x - x_min] + 2)
                write_ws.cell(idx, 5, TX)
                write_ws.cell(idx, 6, TY)
                write_ws.cell(idx, 7, dim_2[int(TY) - y_min][int(TX) - x_min] + 5)
                try:
                    Tx = [int(TX), int(TY), dim_2[int(TY) - y_min][int(TX) - x_min] + 5]
                except:
                    print("건물위가 아닙니다")
                    exit()
                azimuth = int(azimuth)
                downtilt = int(downtilt)
                Rx = [x, y, dim_1[y - y_min][x - x_min] + 2]
                res = angle_diff(Tx, Rx, azimuth, downtilt)
                # print(res)
                write_ws.cell(idx, 8, res[1])
                write_ws.cell(idx, 9, res[2])
                write_ws.cell(idx, 10, res[3])
                idx += 1

    # Grid 좌표 엑셀 파일 이름 설정해줘야함
    write_wb.save("project/static/excel/" + id + ".xlsx")
    result = os.popen("python project/lams_on_svr.py grid " + id).read()
    print(result)


if __name__ == "__main__":
    main()
